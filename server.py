#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Serveur local - Gestionnaire de Projets
Sert index.html, styles.css, app.js et l'API REST
"""

import http.server
import http.client
import email
import email.utils
import email.message
import json
import io
import datetime
import copy
import os
import sys
import webbrowser
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    EXCEL_OK = True
except ImportError:
    EXCEL_OK = False
from pathlib import Path
from urllib.parse import urlparse, parse_qs

# BASE_DIR dynamique : dossier du .exe en prod, dossier du script en dev
if getattr(sys, 'frozen', False):
    BASE_DIR = Path(sys.executable).parent
else:
    BASE_DIR = Path(__file__).parent

DATA_DIR    = BASE_DIR / "data"
LEGACY_FILE = BASE_DIR / "projets.json"
ARCHIVE_DIR = BASE_DIR / "archives"
HIST_FILE   = BASE_DIR / "historique.json"
TUTO_FILE   = BASE_DIR / "tuto.txt"
SETTINGS_FILE = BASE_DIR / "settings.json"
MAX_VERSIONS = 5
PORT         = 8742

STATIC_FILES = {
    "/":                ("index.html",       "text/html; charset=utf-8"),
    "/index.html":      ("index.html",       "text/html; charset=utf-8"),
    "/styles.css":      ("styles.css",       "text/css; charset=utf-8"),
    "/app.js":          ("app.js",           "application/javascript; charset=utf-8"),
    "/aide.html":       ("aide.html",        "text/html; charset=utf-8"),
    "/sw.js":           ("sw.js",            "application/javascript; charset=utf-8"),
    "/favicon.svg":     ("favicon.svg",      "image/svg+xml"),
    "/morphdom.min.js": ("morphdom.min.js",  "application/javascript"),
    "/presentation":    ("presentation.html","text/html; charset=utf-8"),
    "/presentation.html":("presentation.html","text/html; charset=utf-8"),
}

def read_json(path):
    """Lit un fichier JSON en gerant le BOM UTF-8."""
    return json.loads(Path(path).read_text(encoding='utf-8-sig'))

def write_json(path, data):
    """Ecrit un fichier JSON sans BOM."""
    Path(path).write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding='utf-8')

def proj_dirname(p):
    import re
    name = p.get('name', p['id'])
    name = re.sub(r'[<>:"/\\|?*]', '_', name)
    name = name.strip('. ')[:60]
    return name or p['id']

def find_proj_dir(pid):
    if not DATA_DIR.exists():
        return None
    for d in DATA_DIR.iterdir():
        if not d.is_dir(): continue
        pf = d / 'projet.json'
        if pf.exists():
            try:
                p = read_json(pf)
                if p.get('id') == pid:
                    return d
            except: pass
    return None

def ensure_data_dir():
    DATA_DIR.mkdir(parents=True, exist_ok=True)

def load_history():
    if HIST_FILE.exists():
        try: return read_json(HIST_FILE)
        except: return []
    return []

def save_snapshot(action='modification'):
    try:
        data = load_data()
        history = load_history()
        snapshot = {
            'date':   datetime.datetime.now().strftime('%d/%m/%Y %H:%M:%S'),
            'action': action,
            'data':   data
        }
        history.insert(0, snapshot)
        history = history[:MAX_VERSIONS]
        write_json(HIST_FILE, history)
    except Exception as e:
        print(f"Erreur snapshot: {e}")

def load_data():
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    projects, tasks = [], {}
    for proj_dir in sorted(DATA_DIR.iterdir(), key=lambda e: e.name):
        if not proj_dir.is_dir(): continue
        pf = proj_dir / 'projet.json'
        tf = proj_dir / 'taches.json'
        if not pf.exists(): continue
        try:
            p = read_json(pf)
            projects.append(p)
            tasks[p['id']] = read_json(tf) if tf.exists() else []
        except Exception as e:
            print(f"Erreur lecture {proj_dir.name}: {e}")
    # Projets archives
    ARCHIVE_DIR.mkdir(exist_ok=True)
    for zip_path in sorted(ARCHIVE_DIR.iterdir()):
        if zip_path.suffix != '.zip': continue
        try:
            import zipfile
            with zipfile.ZipFile(zip_path, 'r') as z:
                names = z.namelist()
                proj_files  = [n for n in names if n.endswith('projet.json')]
                tasks_files = [n for n in names if n.endswith('taches.json')]
                if proj_files:
                    p = json.loads(z.read(proj_files[0]).decode('utf-8-sig'))
                    p['archived'] = True
                    projects.append(p)
                    tasks[p['id']] = json.loads(z.read(tasks_files[0]).decode('utf-8-sig')) if tasks_files else []
        except Exception as e:
            print(f"Erreur archive {zip_path.name}: {e}")
    return {"projects": projects, "tasks": tasks}

def save_data(data):
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    projects, tasks = data.get('projects', []), data.get('tasks', {})
    existing_ids = {p['id'] for p in projects}
    # Supprimer les dossiers orphelins
    for entry in DATA_DIR.iterdir():
        if not entry.is_dir(): continue
        pjson = entry / 'projet.json'
        if not pjson.exists(): continue
        try:
            p = read_json(pjson)
            if p.get('id') not in existing_ids:
                import shutil; shutil.rmtree(entry)
        except: pass
    for p in projects:
        if p.get('archived'): continue
        existing = find_proj_dir(p['id'])
        new_name = proj_dirname(p)
        new_path = DATA_DIR / new_name
        if existing and existing.name != new_name:
            existing.rename(new_path); pdir = new_path
        elif existing:
            pdir = existing
        else:
            pdir = new_path
        pdir.mkdir(exist_ok=True)
        for sub in ['images', 'stl', 'cura', 'docs']:
            (pdir / sub).mkdir(exist_ok=True)
        write_json(pdir / 'projet.json', p)
        write_json(pdir / 'taches.json', tasks.get(p['id'], []))
    try: save_snapshot()
    except Exception as e: print(f"Erreur snapshot: {e}")


class Handler(http.server.BaseHTTPRequestHandler):

    def log_message(self, format, *args): pass

    def _cors(self):
        self.send_header("Access-Control-Allow-Origin",  "*")
        self.send_header("Access-Control-Allow-Methods", "GET, POST, DELETE, OPTIONS")
        self.send_header("Access-Control-Allow-Headers", "Content-Type")

    def _send_json(self, data, code=200):
        body = json.dumps(data, ensure_ascii=False).encode("utf-8")
        self.send_response(code)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Content-Length", str(len(body)))
        self._cors()
        self.end_headers()
        self.wfile.write(body)

    def _send_file(self, filename, content_type):
        filepath = BASE_DIR / filename
        if not filepath.exists():
            self.send_response(404); self.end_headers(); return
        content = filepath.read_bytes()
        self.send_response(200)
        self.send_header("Content-Type", content_type)
        self.send_header("Content-Length", str(len(content)))
        self.send_header("Cache-Control", "no-cache")
        self._cors()
        self.end_headers()
        self.wfile.write(content)

    def _read_body(self):
        length = int(self.headers.get("Content-Length", 0))
        return json.loads(self.rfile.read(length).decode("utf-8")) if length else {}

    def do_OPTIONS(self):
        self.send_response(200); self._cors(); self.end_headers()

    def do_HEAD(self):
        self.send_response(200)
        self.send_header('Content-Type', 'application/json')
        self.end_headers()

    def do_GET(self):
        path = urlparse(self.path).path
        qs   = parse_qs(urlparse(self.path).query)

        if path in STATIC_FILES:
            filename, content_type = STATIC_FILES[path]
            self._send_file(filename, content_type)

        elif path == "/api/data":
            self._send_json(load_data())

        elif path == "/api/archives":
            ARCHIVE_DIR.mkdir(exist_ok=True)
            archives = [{'name': f.stem, 'file': f.name,
                'size': f.stat().st_size,
                'date': datetime.datetime.fromtimestamp(f.stat().st_mtime).strftime('%d/%m/%Y %H:%M')}
                for f in sorted(ARCHIVE_DIR.iterdir()) if f.suffix == '.zip']
            self._send_json({'archives': archives})

        elif path == "/api/history":
            self._send_json(load_history())

        elif path.startswith("/api/restore"):
            idx = int(qs.get('idx', [0])[0])
            history = load_history()
            if 0 <= idx < len(history):
                save_data(history[idx]['data'])
                self._send_json({'ok': True, 'date': history[idx]['date']})
            else:
                self._send_json({'ok': False, 'error': 'Index invalide'})

        elif path.startswith("/api/files/list"):
            proj_id = qs.get('projId', [None])[0]
            subdir  = qs.get('dir', ['images'])[0]
            if not proj_id: self._send_json({'error': 'projId manquant'}); return
            pdir = find_proj_dir(proj_id)
            if not pdir: self._send_json({'files': []}); return
            target = pdir / subdir; target.mkdir(exist_ok=True)
            files = [{'name': f.name, 'size': f.stat().st_size,
                'url': f'/api/files/get?projId={proj_id}&dir={subdir}&file={f.name}'}
                for f in sorted(target.iterdir()) if f.is_file()]
            self._send_json({'files': files})

        elif path.startswith("/api/files/get"):
            proj_id = qs.get('projId', [None])[0]
            subdir  = qs.get('dir',    [None])[0]
            fname   = qs.get('file',   [None])[0]
            pdir = find_proj_dir(proj_id) if proj_id else None
            if not pdir or not subdir or not fname:
                self.send_response(404); self.end_headers(); return
            fpath = pdir / subdir / fname
            if not fpath.exists(): self.send_response(404); self.end_headers(); return
            import mimetypes
            ctype = mimetypes.guess_type(str(fpath))[0] or 'application/octet-stream'
            content = fpath.read_bytes()
            self.send_response(200)
            self.send_header('Content-Type', ctype)
            self.send_header('Content-Length', str(len(content)))
            self.send_header('Content-Disposition', f'inline; filename="{fname}"')
            self._cors(); self.end_headers(); self.wfile.write(content)

        elif path == "/api/tuto":
            done = TUTO_FILE.exists() and TUTO_FILE.read_text().strip() == '1'
            self._send_json({'done': done})

        elif path == "/api/settings":
            key = qs.get('key', [None])[0]
            settings = read_json(SETTINGS_FILE) if SETTINGS_FILE.exists() else {}
            if key:
                self._send_json({'value': settings.get(key, '')})
            else:
                self._send_json(settings)

        elif path.startswith("/api/export-excel"):
            self._export_excel()

        else:
            self.send_response(404); self.end_headers()

    def do_POST(self):
        parsed = urlparse(self.path)

        if parsed.path == "/api/archive":
            self._archive_project(); return
        if parsed.path == "/api/unarchive":
            self._unarchive_project(); return
        if parsed.path == "/api/delete-archive":
            self._delete_archive(); return
        if parsed.path == "/api/tasks/reorder":
            body = self._read_body()
            pid  = body.get("projectId")
            tks  = body.get("tasks", [])
            if pid:
                pdir = find_proj_dir(pid)
                if pdir: write_json(pdir / "taches.json", tks)
            self._send_json({"ok": True}); return
        if parsed.path == "/api/projects/reorder":
            body = self._read_body()
            ids  = body.get("ids", [])
            d    = load_data()
            pmap = {p['id']: p for p in d['projects']}
            d['projects'] = [pmap[i] for i in ids if i in pmap]
            save_data(d)
            self._send_json({"ok": True}); return
        if parsed.path == "/api/files/upload":
            self._upload_file(); return
        if parsed.path == "/api/files/delete":
            self._delete_file(); return
        if parsed.path == "/api/tuto":
            body = self._read_body()
            action = body.get('action')
            if action == 'done':  TUTO_FILE.write_text('1')
            elif action == 'reset': TUTO_FILE.write_text('0')
            self._send_json({'ok': True}); return
        if parsed.path == "/api/settings":
            body = self._read_body()
            key  = body.get('key')
            val  = body.get('value')
            settings = read_json(SETTINGS_FILE) if SETTINGS_FILE.exists() else {}
            if key:
                if val is not None: settings[key] = val
                else: settings.pop(key, None)
                write_json(SETTINGS_FILE, settings)
            self._send_json({'ok': True}); return

        body = self._read_body()
        data = load_data()

        if parsed.path == "/api/projects":
            proj = body
            existing = next((i for i, p in enumerate(data["projects"]) if p["id"] == proj["id"]), None)
            if existing is not None:
                data["projects"][existing] = proj
            else:
                data["projects"].append(proj)
                if proj["id"] not in data["tasks"]:
                    data["tasks"][proj["id"]] = []
            save_data(data)
            self._send_json({"ok": True})

        elif parsed.path == "/api/tasks":
            pid  = body.get("projectId")
            task = body.get("task")
            if pid and task:
                if pid not in data["tasks"]: data["tasks"][pid] = []
                existing = next((i for i, t in enumerate(data["tasks"][pid]) if t["id"] == task["id"]), None)
                if existing is not None: data["tasks"][pid][existing] = task
                else: data["tasks"][pid].append(task)
                save_data(data)
                self._send_json({"ok": True})
            else:
                self._send_json({"error": "missing data"}, 400)

        else:
            self._send_json({"error": "not found"}, 404)

    def do_DELETE(self):
        parsed = urlparse(self.path)
        qs     = parse_qs(parsed.query)
        data   = load_data()

        if parsed.path == "/api/projects":
            pid = qs.get("id", [None])[0]
            if pid:
                pdir = find_proj_dir(pid)
                if pdir and pdir.exists():
                    import shutil; shutil.rmtree(pdir)
                data["projects"] = [p for p in data["projects"] if p["id"] != pid]
                data["tasks"].pop(pid, None)
                save_data(data)
                self._send_json({"ok": True})

        elif parsed.path == "/api/tasks":
            pid = qs.get("projectId", [None])[0]
            tid = qs.get("taskId",    [None])[0]
            if pid and tid:
                data["tasks"][pid] = [t for t in data["tasks"].get(pid, []) if t["id"] != tid]
                save_data(data)
                self._send_json({"ok": True})
        else:
            self._send_json({"error": "not found"}, 404)

    def _archive_project(self):
        qs      = parse_qs(urlparse(self.path).query)
        proj_id = qs.get('projId', [None])[0]
        if not proj_id: self._send_json({'error': 'projId manquant'}); return
        pdir = find_proj_dir(proj_id)
        if not pdir: self._send_json({'error': 'Projet introuvable'}); return
        ARCHIVE_DIR.mkdir(exist_ok=True)
        import zipfile
        zip_path = ARCHIVE_DIR / (pdir.name + '.zip')
        with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as z:
            for f in pdir.rglob('*'):
                if f.is_file(): z.write(f, f.relative_to(pdir.parent))
        data = load_data()
        for p in data['projects']:
            if p['id'] == proj_id: p['archived'] = True; break
        save_data(data)
        import shutil
        if pdir.exists(): shutil.rmtree(pdir)
        self._send_json({'ok': True, 'zip': zip_path.name, 'size': zip_path.stat().st_size})

    def _unarchive_project(self):
        qs      = parse_qs(urlparse(self.path).query)
        proj_id = qs.get('projId', [None])[0]
        if not proj_id: self._send_json({'error': 'projId manquant'}); return
        data = load_data()
        proj = next((p for p in data['projects'] if p['id'] == proj_id), None)
        if not proj: self._send_json({'error': 'Projet introuvable'}); return
        import re as _re, zipfile
        zip_name = _re.sub(r'[<>:"/\\|?*]', '_', proj.get('name', proj_id)).strip('. ')[:60] + '.zip'
        zip_path = ARCHIVE_DIR / zip_name
        if zip_path.exists():
            with zipfile.ZipFile(zip_path, 'r') as z: z.extractall(DATA_DIR)
            zip_path.unlink()
        proj['archived'] = False
        save_data(data)
        self._send_json({'ok': True})

    def _delete_archive(self):
        qs      = parse_qs(urlparse(self.path).query)
        proj_id = qs.get('projId', [None])[0]
        if not proj_id: self._send_json({'error': 'projId manquant'}); return
        data = load_data()
        proj = next((p for p in data['projects'] if p['id'] == proj_id), None)
        if not proj: self._send_json({'error': 'Projet introuvable'}); return
        import re as _re
        zip_name = _re.sub(r'[<>:"/\\|?*]', '_', proj.get('name', proj_id)).strip('. ')[:60] + '.zip'
        zip_path = ARCHIVE_DIR / zip_name
        if zip_path.exists(): zip_path.unlink()
        data['projects'] = [p for p in data['projects'] if p['id'] != proj_id]
        data['tasks'].pop(proj_id, None)
        save_data(data)
        self._send_json({'ok': True})

    def _upload_file(self):
        qs      = parse_qs(urlparse(self.path).query)
        proj_id = qs.get('projId', [None])[0]
        if not proj_id: self._send_json({'error': 'projId manquant'}); return
        pdir = find_proj_dir(proj_id)
        if not pdir: self._send_json({'error': 'Projet introuvable'}); return
        ext_map = {'.stl':'stl','.obj':'stl','.3mf':'stl','.step':'stl','.stp':'stl',
            '.gcode':'cura','.curaproject':'cura','.cura':'cura',
            '.pdf':'docs','.txt':'docs','.doc':'docs','.docx':'docs',
            '.xls':'docs','.xlsx':'docs','.ino':'docs','.py':'docs',
            '.jpg':'images','.jpeg':'images','.png':'images',
            '.gif':'images','.bmp':'images','.webp':'images','.svg':'images'}
        ctype  = self.headers.get('Content-Type', '')
        length = int(self.headers.get('Content-Length', 0))
        body   = self.rfile.read(length)
        boundary = next((p.strip()[9:].strip('"') for p in ctype.split(';') if p.strip().startswith('boundary=')), None)
        if not boundary: self._send_json({'error': 'Pas de boundary'}); return
        bnd   = ('--' + boundary).encode()
        parts = body.split(bnd)
        saved = []
        import re as _re
        for part in parts[1:]:
            if part.strip() in [b'', b'--', b'--\r\n']: continue
            sep = part.find(b'\r\n\r\n')
            if sep < 0: continue
            hdr_raw = part[:sep].decode('utf-8', errors='replace')
            content = part[sep+4:]
            if content.endswith(b'\r\n'): content = content[:-2]
            m = _re.search(r'filename="([^"]+)"', hdr_raw)
            if not m: continue
            fname   = m.group(1)
            ext     = Path(fname).suffix.lower()
            autodir = ext_map.get(ext, 'docs')
            target  = pdir / autodir; target.mkdir(exist_ok=True)
            (target / fname).write_bytes(content)
            saved.append({'name': fname, 'size': len(content),
                'url': f'/api/files/get?projId={proj_id}&dir={autodir}&file={fname}'})
        self._send_json({'ok': True, 'files': saved})

    def _delete_file(self):
        qs      = parse_qs(urlparse(self.path).query)
        proj_id = qs.get('projId', [None])[0]
        subdir  = qs.get('dir',    [None])[0]
        fname   = qs.get('file',   [None])[0]
        pdir = find_proj_dir(proj_id) if proj_id else None
        if not pdir or not subdir or not fname:
            self._send_json({'error': 'Parametres manquants'}); return
        fpath = pdir / subdir / fname
        if fpath.exists(): fpath.unlink(); self._send_json({'ok': True})
        else: self._send_json({'error': 'Fichier introuvable'})

    def _export_excel(self):
        qs      = parse_qs(urlparse(self.path).query)
        proj_id = qs.get('projId', [None])[0]
        if not EXCEL_OK:
            self.send_response(500); self.end_headers()
            self.wfile.write(b"openpyxl non installe"); return
        data = load_data()
        projects = [p for p in data.get('projects', []) if (not proj_id or p['id'] == proj_id)]
        tasks    = data.get('tasks', {})
        STATUS = {'todo':'A faire','prog':'En cours','done':'Termine','block':'Bloque'}
        PRIO   = {'high':'Haute','med':'Moyenne','low':'Basse'}
        hf = Font(bold=True, color='FFFFFF', size=11)
        hfill = PatternFill('solid', fgColor='0073EA')
        ha = Alignment(horizontal='center', vertical='center')
        bd = Side(style='thin', color='E6E9EF')
        bdr = Border(left=bd, right=bd, top=bd, bottom=bd)
        wb = openpyxl.Workbook(); wb.remove(wb.active)
        ws = wb.create_sheet("Projets")
        for c, h in enumerate(['Projet','Type','Statut','Priorite','Deadline','Taches','Budget est.','Budget reel'], 1):
            cell = ws.cell(row=1, column=c, value=h)
            cell.font=hf; cell.fill=hfill; cell.alignment=ha; cell.border=bdr
        for ri, p in enumerate(projects, 2):
            pt = tasks.get(p['id'], [])
            for ci, val in enumerate([p.get('name',''), p.get('type',''),
                STATUS.get(p.get('status',''),''), PRIO.get(p.get('priority',''),''),
                p.get('deadline',''), len(pt),
                p.get('budgetEst', 0) or 0, p.get('budgetReal', 0) or 0], 1):
                cell = ws.cell(row=ri, column=ci, value=val); cell.border=bdr
                if ri % 2 == 0: cell.fill = PatternFill('solid', fgColor='F6F7FB')
        output = io.BytesIO(); wb.save(output); xlsx = output.getvalue()
        fname = f"projet_{proj_id[:8]}.xlsx" if proj_id else "projets_export.xlsx"
        self.send_response(200)
        self.send_header('Content-Type','application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        self.send_header('Content-Disposition', f'attachment; filename="{fname}"')
        self.send_header('Content-Length', str(len(xlsx)))
        self._cors(); self.end_headers(); self.wfile.write(xlsx)


if __name__ == "__main__":
    ensure_data_dir()
    print("")
    print("  ========================================")
    print("   Gestionnaire de Projets - Demarrage")
    print("  ========================================")
    print("")
    print(f"  Adresse : http://localhost:{PORT}")
    print(f"  Donnees : {DATA_DIR}")
    print("")
    print("  Ferme cette fenetre pour arreter le serveur")
    print("")
    server = http.server.HTTPServer(("localhost", PORT), Handler)
    webbrowser.open(f"http://localhost:{PORT}")
    try:
        server.serve_forever()
    except KeyboardInterrupt:
        print("\n  Serveur arrete.")
